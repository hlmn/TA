import json
import os
import csv
from openpyxl import Workbook
import requests
import sys
pubsub = {}
dir_path = os.path.dirname(os.path.realpath(__file__))
path = "/Users/muhammadhilman/hlmn/Project/MMT-SmartCampus/Latihan/server"
files = [i for i in os.listdir(path) if os.path.isfile(os.path.join(path,i)) and 'benchmarkPubSub' in i]
for file in files:
    ea = open(path+"/"+file, 'r')
    kelas = file.replace('benchmarkPubSub','')
    kelas = kelas.replace('.txt', '')
    for line in ea:
        line = line.replace('\r', '')
        line = line.replace('\n', '')
        line = line.split('|')
        if line[0] != '':
            if kelas not in pubsub:
                pubsub[kelas] = {}
            pubsub[kelas][line[0]] = {
                    'waktu' : line[1]
                }


# ea = open(dir_path+"/benchmark.txt", 'r')
# print(pubsub)
# check = ['208', '214', '207', '215']
# print(pubsub.keys())
# exit()
r = requests.get('http://localhost:9999/get/ruangan')
# print(r.text)
result = json.loads(r.text)
wbSemua = Workbook()
wsSemua = wbSemua.active
wsSemua.append(['nama ruangan', 'log/s', 'Mean', 'Jumlah Log', 'Total Waktu'])
gagal = 0
logGagal = []
for namaRuangan in result['ruangan']:
    
    jumlahLog = 0
    totalWaktuSemua = 0
    wb = Workbook()
    ws = wb.active
    ws.append(['waktu', 'table'])
    ea = open(dir_path+"/benchmark.txt", 'r')
    for line in ea:
        # print(line)
        line = line.replace('\r\n', '')
        line = line.split('|')
        tujuan = json.loads(line[2])
        for ruangan in tujuan:
            # print(pubsub[ruangan])
            
            if ruangan not in pubsub:
                continue
            if line[0] not in pubsub[ruangan]:
                gagal += 1
                logGagal.append(line[0]+'|'+line[2])
            else: 
                if ruangan == namaRuangan:
                    # print(ruangan)
                    waktuPublish = float(line[1])
                    waktuDeliv = float(pubsub[ruangan][line[0]]['waktu'])
                    log = json.loads(line[0])
                    totalWaktu = (waktuDeliv+waktuPublish)
                    # print(log['table']+' -> '+str(totalWaktu))
                    ws.append([totalWaktu, log['table']])
                    jumlahLog += 1
                    totalWaktuSemua += totalWaktu
                else:
                    continue

                
                # print (str(waktuPublish)+' + '+str(waktuDeliv)+' = '+str(totalWaktu))
                # print pubsub[ruangan][line[0]]

    ws['D1'] = ('log/s')
    ws['E1'] = (jumlahLog/totalWaktuSemua)
    ws['D2'] = ('Mean')
    ws['E2'] = (totalWaktuSemua/jumlahLog)
    ws['D3'] = ('Jumlah Log')
    ws['E3'] = (jumlahLog)
    ws['D4'] = ('Total Waktu')
    ws['E4'] = (totalWaktuSemua)
    print('gagal : ' + str(gagal))

    wsSemua.append([namaRuangan, jumlahLog/totalWaktuSemua, totalWaktuSemua/jumlahLog, jumlahLog, totalWaktuSemua])
    
    wb.save("benchmark/test"+namaRuangan+"-"+sys.argv[1]+".xlsx")


print(gagal)
wsSemua.append([sys.argv[2]])
wbSemua.save("benchmark/Kesimpulan-"+sys.argv[1]+".xlsx")
print(logGagal)


for file in files:
    os.remove(path+"/"+file)
os.rename(dir_path+'/benchmark.txt', dir_path+'/benchmark/benchmark'+sys.argv[1]+'.txt')
# print(dir_path)